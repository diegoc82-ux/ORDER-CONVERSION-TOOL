import io
import time
import json as _json
from pathlib import Path
from flask import Flask, request, jsonify, send_file, render_template
import pdfplumber
from qb_queue import init_db, enqueue, get_job, get_recent_jobs

app = Flask(__name__)
BASE = Path(__file__).parent
init_db()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/products.json')
def products():
    return send_file(BASE / 'products.json', mimetype='application/json')


@app.route('/imporcarsa_prices.json')
def imporcarsa_prices():
    return send_file(BASE / 'imporcarsa_prices.json', mimetype='application/json')


@app.route('/8a_prices.json')
def a8_prices():
    return send_file(BASE / '8a_prices.json', mimetype='application/json')


@app.route('/competitive_edge_prices.json')
def competitive_edge_prices():
    return send_file(BASE / 'competitive_edge_prices.json', mimetype='application/json')


@app.route('/beicruz_prices.json')
def beicruz_prices():
    return send_file(BASE / 'beicruz_prices.json', mimetype='application/json')


@app.route('/garner_prices.json')
def garner_prices():
    return send_file(BASE / 'garner_prices.json', mimetype='application/json')


@app.route('/pagsa_prices.json')
def pagsa_prices():
    return send_file(BASE / 'pagsa_prices.json', mimetype='application/json')


@app.route('/trebol_prices.json')
def trebol_prices():
    return send_file(BASE / 'trebol_prices.json', mimetype='application/json')


@app.route('/maxilub_prices.json')
def maxilub_prices():
    return send_file(BASE / 'maxilub_prices.json', mimetype='application/json')


@app.route('/servistar_prices.json')
def servistar_prices():
    return send_file(BASE / 'servistar_prices.json', mimetype='application/json')


@app.route('/frenoseguro_prices.json')
def frenoseguro_prices():
    return send_file(BASE / 'frenoseguro_prices.json', mimetype='application/json')


@app.route('/ukr_prices.json')
def ukr_prices():
    return send_file(BASE / 'ukr_prices.json', mimetype='application/json')


@app.route('/prosupply_prices.json')
def prosupply_prices():
    return send_file(BASE / 'prosupply_prices.json', mimetype='application/json')


@app.route('/abc_prices.json')
def abc_prices():
    return send_file(BASE / 'abc_prices.json', mimetype='application/json')


@app.route('/daher_prices.json')
def daher_prices():
    return send_file(BASE / 'daher_prices.json', mimetype='application/json')


@app.route('/asaray_prices.json')
def asaray_prices():
    return send_file(BASE / 'asaray_prices.json', mimetype='application/json')


@app.route('/dianca_prices.json')
def dianca_prices():
    return send_file(BASE / 'dianca_prices.json', mimetype='application/json')


@app.route('/costs.json')
def costs():
    return send_file(BASE / 'costs.json', mimetype='application/json')


@app.route('/houston_prices.json')
def houston_prices():
    return send_file(BASE / 'houston_prices.json', mimetype='application/json')


@app.route('/miami_prices.json')
def miami_prices():
    return send_file(BASE / 'miami_prices.json', mimetype='application/json')


@app.route('/forwarders.json')
def forwarders_list():
    p = BASE / 'forwarders.json'
    if not p.exists():
        return jsonify([])
    return send_file(p, mimetype='application/json')


@app.route('/api/save-forwarder', methods=['POST'])
def save_forwarder():
    import json as json_mod
    data = request.get_json()
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Name required'}), 400
    p = BASE / 'forwarders.json'
    existing = json_mod.loads(p.read_text('utf-8')) if p.exists() else []
    if any(f['name'].upper() == name.upper() for f in existing):
        return jsonify({'error': 'Forwarder already exists'}), 409
    fwd_id = name.lower().replace(' ', '_').replace('-', '_')
    existing.append({
        'id':       fwd_id,
        'name':     name,
        'address1': (data.get('address1') or '').strip(),
        'address2': (data.get('address2') or '').strip(),
        'address3': (data.get('address3') or '').strip(),
    })
    p.write_text(json_mod.dumps(existing, indent=2, ensure_ascii=False), encoding='utf-8')
    return jsonify({'ok': True, 'id': fwd_id})


@app.route('/api/extract-dispatch', methods=['POST'])
def extract_dispatch():
    import re
    from collections import defaultdict
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    try:
        items, dispatch_num, po_number, order_date = [], '', '', ''
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                # Extract dispatch number ("Dispatch Note XXXX")
                for i, w in enumerate(words):
                    if w['text'] == 'Note' and i > 0 and words[i-1]['text'] == 'Dispatch':
                        if i+1 < len(words):
                            dispatch_num = words[i+1]['text']
                # Extract PO number and order date from right-side header
                for w in words:
                    if w['text'] == 'Number:' and w['x0'] > 500:
                        val = next((v['text'] for v in words
                                    if v['x0'] > 650 and abs(v['top'] - w['top']) < 5), '')
                        po_number = val
                    if w['text'] == 'order:' and w['x0'] > 500:
                        val = next((v['text'] for v in words
                                    if v['x0'] > 650 and abs(v['top'] - w['top']) < 5), '')
                        order_date = val
                # Group words by rounded top value (2px tolerance for sub-pixel differences)
                rows = defaultdict(list)
                for w in words:
                    rows[round(w['top'] / 2) * 2].append(w)
                # Find header row and detect column positions.
                # Supports two formats:
                #   Standard U1P:   CODE … SPEC … UNIT QTY …
                #   FRENOSEGURO:    UL Number … PACKAGE SIZE … UNIT QTY …
                header_top = None
                spec_lo = spec_hi = qty_lo = qty_hi = None
                for top, rw in sorted(rows.items()):
                    has_code = any(w['text'] == 'CODE' for w in rw)
                    has_spec = any(w['text'] == 'SPEC' for w in rw)
                    has_number = any(w['text'] == 'Number' for w in rw)
                    has_package = any(w['text'] == 'PACKAGE' for w in rw)
                    if (has_code and has_spec) or (has_number and has_package):
                        header_top = top
                        # Anchor for spec column: SPEC (standard) or PACKAGE (FRENOSEGURO)
                        spec_anchor = (next((w for w in rw if w['text'] == 'SPEC'), None)
                                       or next((w for w in rw if w['text'] == 'PACKAGE'), None))
                        unit_w = next((w for w in rw if w['text'] == 'UNIT'), None)
                        qty_candidates = sorted(
                            [w for w in rw if w['text'] == 'QTY' and unit_w and w['x0'] > unit_w['x1']],
                            key=lambda x: x['x0']
                        )
                        unit_qty_w = qty_candidates[0] if qty_candidates else None
                        if spec_anchor and unit_w:
                            spec_lo = spec_anchor['x0'] - 15
                            spec_hi = unit_w['x0'] - 3
                            qty_lo  = unit_w['x0'] - 5
                            qty_hi  = (unit_qty_w['x1'] if unit_qty_w else unit_w['x1']) + 35
                        break
                if header_top is None or spec_lo is None:
                    continue
                # Extract line items from rows below the header
                for top, row_words in sorted(rows.items()):
                    if top <= header_top:
                        continue
                    code_words = [w for w in row_words
                                  if re.match(r'^(UL|TR)\d+$', w['text'], re.I) and w['x0'] < 130]
                    if not code_words:
                        continue
                    code = code_words[0]['text'].upper()
                    spec_words = sorted([w for w in row_words if spec_lo <= w['x0'] <= spec_hi],
                                        key=lambda x: x['x0'])
                    presentation = ' '.join(w['text'] for w in spec_words)
                    if not presentation:
                        continue
                    qty_words = [w for w in row_words if qty_lo <= w['x0'] <= qty_hi]
                    if not qty_words:
                        continue
                    try:
                        qty = int(float(qty_words[0]['text'].replace(',', '')))
                    except (ValueError, TypeError):
                        continue
                    if qty <= 0:
                        continue
                    items.append({'ul_code': code, 'presentation': presentation, 'qty': qty})
        return jsonify({'items': items, 'dispatchNum': dispatch_num,
                        'poNumber': po_number, 'orderDate': order_date})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/extract-garner', methods=['POST'])
def extract_garner():
    import re
    from collections import defaultdict
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    try:
        # Column x-midpoint ranges → presentation
        COL_PRES = [
            (690, 730, 'BOX (6G)'),    # 6/1-Gal Case
            (630, 690, 'PAIL (5G)'),   # 5G Pail
            (570, 630, 'DRUM (55G)'),  # 55G Drum
            (525, 570, 'TOTE (250G)'), # Tote (coolant bulk)
            (460, 525, 'BULK (COOL)'), # Bulk per gallon
        ]
        items, po_number, order_date = [], '', ''
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                # Extract PO number and date
                for i, w in enumerate(words):
                    if re.match(r'^PO#$', w['text'], re.I):
                        # PO# and its number are separate tokens — grab the next word
                        if i + 1 < len(words) and re.match(r'^\d+$', words[i + 1]['text']):
                            po_number = 'PO#' + words[i + 1]['text']
                        else:
                            po_number = w['text']
                    if re.match(r'^\d{1,2}-\d{1,2}-\d{2,4}$', w['text']):
                        order_date = w['text']
                # Group words into rows by top position (10pt bands)
                rows = defaultdict(list)
                for w in words:
                    rows[round(w['top'] / 10) * 10].append(w)
                for row_words in rows.values():
                    ul = next((w for w in row_words if re.match(r'^UL\d+$', w['text'])), None)
                    if not ul:
                        continue
                    for w in row_words:
                        try:
                            qty = float(w['text'].replace(',', ''))
                        except (ValueError, TypeError):
                            continue
                        if qty <= 0:
                            continue
                        x_mid = (w['x0'] + w['x1']) / 2
                        pres = next((p for lo, hi, p in COL_PRES if lo <= x_mid <= hi), None)
                        if pres:
                            items.append({'ul_code': ul['text'], 'presentation': pres, 'qty': int(qty)})
        return jsonify({'items': items, 'poNumber': po_number, 'orderDate': order_date})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/extract-ukr', methods=['POST'])
def extract_ukr():
    import re
    from collections import defaultdict
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    try:
        items, po_number, order_date = [], '', ''
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                # Extract PO number: "PURCHASE ORDER # 74"
                for i, w in enumerate(words):
                    if w['text'] == '#' and i >= 2 and words[i-2]['text'] == 'PURCHASE':
                        if i + 1 < len(words) and words[i+1]['text'].isdigit():
                            po_number = 'PO_' + words[i+1]['text'].zfill(5)
                    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', w['text']):
                        order_date = w['text']
                # Group words by row
                rows = defaultdict(list)
                for w in words:
                    rows[round(w['top'] / 3) * 3].append(w)
                # Find header row (has "SKU" and "Qty" columns)
                header_top = sku_lo = sku_hi = qty_lo = qty_hi = None
                for top, rw in sorted(rows.items()):
                    texts = [w['text'] for w in rw]
                    if 'SKU' in texts and 'Qty' in texts:
                        header_top = top
                        sku_w = next((w for w in rw if w['text'] == 'SKU'), None)
                        qty_w = next((w for w in rw if w['text'] == 'Qty'), None)
                        if sku_w and qty_w:
                            sku_lo = sku_w['x0'] - 5
                            sku_hi = sku_w['x1'] + 130
                            qty_lo = qty_w['x0'] - 10
                            qty_hi = qty_w['x1'] + 35
                        break
                if header_top is None or sku_lo is None:
                    continue
                for top, row_words in sorted(rows.items()):
                    if top <= header_top:
                        continue
                    sku_cands = [w for w in row_words
                                 if sku_lo <= w['x0'] <= sku_hi
                                 and re.match(r'^[A-Z]{2}\w{4,}$', w['text'], re.I)]
                    if not sku_cands:
                        continue
                    sku = sku_cands[0]['text'].upper()
                    qty_cands = [w for w in row_words
                                 if qty_lo <= w['x0'] <= qty_hi
                                 and re.match(r'^\d+$', w['text'])]
                    if not qty_cands:
                        continue
                    qty = int(qty_cands[0]['text'])
                    if qty <= 0:
                        continue
                    items.append({'sku': sku, 'qty': qty})
        return jsonify({'items': items, 'poNumber': po_number, 'orderDate': order_date})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/extract-pdf', methods=['POST'])
def extract_pdf():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    try:
        with pdfplumber.open(io.BytesIO(f.read())) as pdf:
            text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        return jsonify({'text': text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/counter', methods=['GET'])
def get_counter():
    import json
    path = BASE / 'dispatch_counter.json'
    return jsonify(json.loads(path.read_text()) if path.exists() else {'last': 3086})


@app.route('/api/counter', methods=['POST'])
def set_counter():
    import json
    path = BASE / 'dispatch_counter.json'
    path.write_text(json.dumps(request.get_json()))
    return jsonify({'ok': True})


@app.route('/api/generate-sli', methods=['POST'])
def generate_sli():
    import re
    import openpyxl
    from datetime import datetime, date
    import io as io_mod

    data        = request.get_json()
    consignee   = data.get('consignee', {})
    agent       = data.get('forwarding_agent', {})
    ref         = data.get('reference', '')
    date_str    = data.get('date', date.today().isoformat())
    items       = data.get('items', [])

    try:
        sli_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        sli_date = date.today()

    GALS = {
        'BOX (12Q)':      3.0,   'BOX (3/5QTS)':   3.75,
        'BOX (4G)':       4.0,   'BOX (6G)':        6.0,
        'BOX (1 X 2.5G)': 2.5,   'BOX (2 X 2.5G)':  5.0,
        'DRUM (55G)':    55.0,   'PAIL (5G)':        5.0,
        'TOTE (265G)':  265.0,   'TOTE (250G)':    250.0,
        'TOTE (330G)':  330.0,   'BULK (OIL)':       1.0,
        'BULK (COOL)':    1.0,   'JERRYCAN (20L)':   5.283,
        'CASE 10/1':      2.5,
    }

    def cat(it):
        c = it.get('ul_code', '')
        if c == 'UL990': return 'def'
        if re.match(r'^UL9\d{2}$', c): return 'coolant'
        return 'lube'

    # Separate discount pseudo-items from real products
    discount_total = sum(it.get('value_usd', 0) for it in items if it.get('ul_code') == '__DISCOUNT__')
    product_items  = [it for it in items if it.get('ul_code') != '__DISCOUNT__']

    groups = {'coolant': [], 'lube': [], 'def': []}
    for it in product_items:
        groups[cat(it)].append(it)

    # For proportional discount distribution across HTS categories
    total_product_value = sum(it.get('value_usd', 0) for grp in groups.values() for it in grp)
    def grp_discount(grp):
        if total_product_value <= 0 or discount_total >= 0: return 0
        grp_val = sum(it.get('value_usd', 0) for it in grp)
        return round(discount_total * grp_val / total_product_value, 2)

    def calc_group(grp, extra_value=0):
        total_gal = sum(it['qty'] * GALS.get(it['presentation'], 0) for it in grp)
        bbl       = total_gal / 42.0
        weight_kg = sum(it.get('weight_lbs', 0) for it in grp) / 2.20462
        value_usd = sum(it.get('value_usd', 0) for it in grp) + extra_value
        boxes = sum(it['qty'] for it in grp if 'BOX' in it.get('presentation', ''))
        drums = sum(it['qty'] for it in grp if 'DRUM' in it.get('presentation', ''))
        pails = sum(it['qty'] for it in grp if 'PAIL' in it.get('presentation', ''))
        totes = sum(it['qty'] for it in grp if 'TOTE' in it.get('presentation', ''))
        bulk  = sum(it['qty'] for it in grp if 'BULK' in it.get('presentation', ''))
        jerry = sum(it['qty'] for it in grp if 'JERRYCAN' in it.get('presentation', ''))
        parts = []
        if boxes: parts.append(f"{boxes} cases")
        if drums: parts.append(f"{drums} drums")
        if pails: parts.append(f"{pails} pails")
        if totes: parts.append(f"{totes} totes")
        if bulk:  parts.append(f"{bulk} gallons (bulk)")
        if jerry: parts.append(f"{jerry} jerrycans")
        return round(bbl, 2), round(weight_kg, 3), round(value_usd, 2), (' and '.join(parts) or '0 units')

    product_lines = []
    if groups['coolant']:
        bbl, wkg, val, desc = calc_group(groups['coolant'], grp_discount(groups['coolant']))
        product_lines.append(('D', '3820000000', f'{bbl:.2f} BBL', desc, wkg, 'EAR99', 'N', 'N/A', val, 'N/A'))
    if groups['lube']:
        bbl, wkg, val, desc = calc_group(groups['lube'], grp_discount(groups['lube']))
        product_lines.append(('D', '2710193020', f'{bbl:.2f} BBL', desc, wkg, 'EAR99', 'N', 'N/A', val, 'N/A'))
    if groups['def']:
        bbl, wkg, val, desc = calc_group(groups['def'], grp_discount(groups['def']))
        product_lines.append(('D', '3102100000', f'{bbl:.2f} BBL', desc, wkg, 'EAR99', 'N', 'N/A', val, 'N/A'))

    total_kg = sum(it.get('weight_lbs', 0) for it in items) / 2.20462

    # ── Load the formatted template and fill in variable fields only ──────────
    template_path = BASE / 'sli_template.xlsx'
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    def w(cell_ref, value):
        """Write a value to a cell (top-left of merged range) without touching formatting."""
        ws[cell_ref] = value

    # ── USPPI — always Ultrachem LLC (boxes 1, 2, 6) ─────────────────────────
    w('A3', 'Ultrachem LLC')
    w('A5', '1444 Northwest 82nd Ave.')
    w('A6', 'Doral, FL 33126')
    w('C8', '82-3520413')          # USPPI EIN (IRS) No

    # ── Freight Location — always U1Dynamics (boxes 3, 4) ────────────────────
    w('E3', 'U1Dynamics Manufacturing')
    w('E5', '4468 Genoa-Red Bluff Rd,')
    w('E6', 'Pasadena, TX. 77505')

    # ── Forwarding Agent (column J, rows 3-6) ────────────────────────────────
    w('J3', agent.get('name', ''))
    w('J4', agent.get('address1', ''))
    w('J5', agent.get('address2', ''))
    w('J6', agent.get('address3', ''))

    # ── Checkbox / select-one fields (static defaults) ───────────────────────
    # Box 7 — Related Party Indicator
    w('J8', '\u2610  Related')
    w('L8', '\u2611  Non-Related')
    # Box 9 — Routed Export Transaction
    w('J9', '\u2610  Yes')
    w('L9', '\u2611  No')
    # Box 11 — Ultimate Consignee Type
    w('E11', '\u2610  Direct Consumer')
    w('E12', '\u2610  Government Entity')
    w('E13', '\u2611  Reseller')
    w('E14', '\u2610  Other/Unknown')
    # Box 15 — Hazardous Material
    w('D18', '\u2610  Yes          \u2611  No')
    # Box 19 — TIB / Carnet
    w('L17', '\u2610  Yes')
    w('L18', '\u2611  No')

    # ── Reference # ──────────────────────────────────────────────────────────
    w('C9', ref)

    # ── Consignee (rows 11-14) ────────────────────────────────────────────────
    tax_id = consignee.get('tax_id', '')
    w('A11', consignee.get('name', ''))
    if tax_id:
        w('A12', tax_id)
        w('A13', consignee.get('address1', ''))
        w('A14', consignee.get('address2', ''))
    else:
        w('A12', consignee.get('address1', ''))
        w('A13', consignee.get('address2', ''))
        w('A14', '')

    # ── Country of Ultimate Destination ──────────────────────────────────────
    w('D17', consignee.get('country', '').upper())

    # ── Gross Weight (kg) ────────────────────────────────────────────────────
    w('B20', round(total_kg, 3))

    # ── Product Lines (rows 22-26) ───────────────────────────────────────────
    # Clear all 5 data rows first
    for r in range(22, 27):
        for col in ['A', 'B', 'D', 'E', 'F', 'G', 'H', 'I', 'L', 'M']:
            ws[f'{col}{r}'] = None

    for i, (df, hts, bbl_str, units_desc, wkg, eccn, sme, lic, val, licval) in enumerate(product_lines):
        r = 22 + i
        w(f'A{r}', df)
        w(f'B{r}', hts)    # B:C merged
        w(f'D{r}', bbl_str)
        w(f'E{r}', units_desc)
        w(f'F{r}', wkg)
        w(f'G{r}', eccn)
        w(f'H{r}', sme)
        w(f'I{r}', lic)    # I:K merged
        w(f'L{r}', val)
        w(f'M{r}', licval)

    # ── Date ─────────────────────────────────────────────────────────────────
    ws['M34'] = sli_date
    ws['M34'].number_format = 'MM/DD/YYYY'

    out = io_mod.BytesIO()
    wb.save(out)
    out.seek(0)
    cn = consignee.get('name', 'EXPORT').replace('/', '-').replace('\\', '-')
    filename = f"SLI - {cn} - {ref}.xlsx"
    return send_file(out, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/notion-lookup-dispatch', methods=['POST'])
def notion_lookup_dispatch():
    import json as _json
    try:
        import requests as _req
    except ImportError:
        return jsonify({'error': 'requests library not installed. Run: pip install requests'}), 500

    data = request.get_json()
    dispatch_num = (data.get('dispatch_num') or '').strip()
    if not dispatch_num:
        return jsonify({'error': 'dispatch_num is required'}), 400

    cfg_path = BASE / 'notion_config.json'
    if not cfg_path.exists():
        return jsonify({'error': 'notion_config.json not found'}), 500
    cfg = _json.loads(cfg_path.read_text())
    token = cfg.get('token', '')
    db_id = cfg.get('database_id', '2dfa8101-da66-811d-89ab-000b9cc2f16c')
    if not token or token.startswith('YOUR_'):
        return jsonify({'error': 'Notion token not configured in notion_config.json'}), 500

    headers = {
        'Authorization': f'Bearer {token}',
        'Notion-Version': '2022-06-28',
        'Content-Type': 'application/json',
    }
    payload = {'filter': {'property': 'title', 'title': {'contains': dispatch_num}}}
    resp = _req.post(f'https://api.notion.com/v1/databases/{db_id}/query',
                     headers=headers, json=payload, timeout=10)
    if resp.status_code != 200:
        return jsonify({'error': f'Notion API error {resp.status_code}: {resp.text[:200]}'}), 500

    results = resp.json().get('results', [])
    if not results:
        return jsonify({'found': False, 'message': f'No record found for "{dispatch_num}"'}), 200

    props = results[0].get('properties', {})

    def get_val(name):
        p = props.get(name, {})
        t = p.get('type', '')
        if t == 'rich_text':
            rt = p.get('rich_text', [])
            return rt[0]['plain_text'] if rt else ''
        if t == 'title':
            ti = p.get('title', [])
            return ti[0]['plain_text'] if ti else ''
        if t == 'select':
            s = p.get('select')
            return s['name'] if s else ''
        if t == 'multi_select':
            return ', '.join(m['name'] for m in p.get('multi_select', []))
        if t == 'number':
            v = p.get('number')
            return str(v) if v is not None else ''
        return ''

    return jsonify({
        'found': True,
        'container_num':      get_val('Container #'),
        'seal_num':           get_val('Seal #'),
        'customer_po':        get_val('CUSTOMER PO'),
        'delivery_location':  get_val('Delivery Location'),
        'pickup_location':    get_val('Pick-Up Location'),
        'booking_bol':        get_val('Booking BOL #'),
        'service_provider':   get_val('Service Provider'),
        'type_of_load':       get_val('Type of Load'),
        'pallet_qty':         get_val('Pallet QTY'),
        'tare_weight':        get_val('Tare Weight'),
    })


@app.route('/api/generate-packing-list', methods=['POST'])
def generate_packing_list():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime, date
    import io as io_mod

    data           = request.get_json()
    items          = [it for it in data.get('items', []) if it.get('ul_code') not in ('__DISCOUNT__', '__SUBTOTAL__')]
    consignee      = data.get('consignee', {})
    reference      = data.get('reference', '')
    po_number      = data.get('po_number', '')
    date_str       = data.get('date', date.today().isoformat())
    container_num  = data.get('container_num', '')
    seal_num       = data.get('seal_num', '')
    port_origin    = data.get('port_origin', '')
    final_dest     = data.get('final_dest', '')
    mode_transport = data.get('mode_transport', '')
    container_wt   = float(data.get('container_weight_lbs', 0) or 0)
    pallets        = int(data.get('pallets', 0) or 0)
    comments       = data.get('comments', 'Proudly made in the USA')

    try:
        doc_date     = datetime.strptime(date_str, '%Y-%m-%d')
        date_display = doc_date.strftime('%B %d, %Y')
    except Exception:
        date_display = date_str

    PIECES_MULT = {
        'BOX (12Q)': 12, 'BOX (4G)': 4,   'BOX (6G)': 6,
        'BOX (3/5QTS)': 3, 'BOX (2 X 2.5G)': 2, 'BOX (1 X 2.5G)': 1,
        'PAIL (5G)': 1, 'DRUM (55G)': 1,   'TOTE (265G)': 1,
        'TOTE (250G)': 1, 'TOTE (330G)': 1, 'JERRYCAN (20L)': 1,
        'CASE 10/1': 1,
    }

    NAVY     = '0D2B4E'
    LT_GRAY  = 'F2F4F6'
    MED_GRAY = 'D8DDE4'
    WHITE    = 'FFFFFF'
    SUM1_BG  = 'E4EBF5'
    SUM2_BG  = 'C8D8EE'
    SUM3_BG  = 'A8BFE0'

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def font(bold=False, color='222222', size=10, italic=False):
        return Font(bold=bold, color=color, size=size, name='Arial', italic=italic)

    def align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def border(left=False, right=False, top=False, bottom=False, thick_bottom=False):
        thin = Side(style='thin', color='BBBBBB')
        thick = Side(style='medium', color=NAVY)
        n = Side(style=None)
        return Border(
            left=thin if left else n,
            right=thin if right else n,
            top=thin if top else n,
            bottom=(thick if thick_bottom else thin) if bottom else n,
        )

    def set_cell(ws, coord, value, fnt=None, fll=None, aln=None, brd=None):
        c = ws[coord]
        c.value = value
        if fnt: c.font = fnt
        if fll: c.fill = fll
        if aln: c.alignment = aln
        if brd: c.border = brd
        return c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Packing List'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 54
    ws.column_dimensions['E'].width = 14

    r = 1

    # ── Logo (top-right, above the date, row 1) ───────────────────────────────
    logo_path = BASE / 'static' / 'u1p_logo.png'
    _logo_img = None
    if logo_path.exists():
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage
            import io as _io
            pil = PILImage.open(logo_path).convert('RGB')
            target_h_px = 52
            ratio = target_h_px / pil.height
            pil = pil.resize((int(pil.width * ratio), target_h_px), PILImage.LANCZOS)
            buf = _io.BytesIO()
            pil.save(buf, format='PNG')
            buf.seek(0)
            _logo_img = XLImage(buf)
            _logo_img.width  = pil.width
            _logo_img.height = target_h_px
            _logo_img.anchor = 'E1'
        except Exception:
            _logo_img = None

    # ── Title row (taller to accommodate logo) ────────────────────────────────
    ws.merge_cells(f'A{r}:D{r}')
    set_cell(ws, f'A{r}', 'PACKING LIST',
             fnt=Font(bold=True, color=WHITE, size=16, name='Arial'),
             fll=fill(NAVY), aln=align('left'))
    # E1 — navy background behind the logo
    set_cell(ws, f'E{r}', '', fll=fill(NAVY))
    ws.row_dimensions[r].height = 58
    r += 1

    # ── Company + Date ────────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:C{r}')
    set_cell(ws, f'A{r}', 'ULTRACHEM LLC',
             fnt=Font(bold=True, color=WHITE, size=11, name='Arial'),
             fll=fill(NAVY), aln=align('left'))
    ws.merge_cells(f'D{r}:E{r}')
    set_cell(ws, f'D{r}', f'Date: {date_display}',
             fnt=font(color=WHITE, size=10), fll=fill(NAVY), aln=align('right'))
    ws.row_dimensions[r].height = 20
    r += 1

    # ── Address row ───────────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:C{r}')
    set_cell(ws, f'A{r}', '1444 NW 82nd Ave. Miami, FL 33126  •  United States of America',
             fnt=Font(color=MED_GRAY, size=8.5, name='Arial'),
             fll=fill(NAVY), aln=align('left'))
    ws.merge_cells(f'D{r}:E{r}')
    set_cell(ws, f'D{r}', 'Ph: 786-953-6132',
             fnt=Font(color=MED_GRAY, size=8.5, name='Arial'),
             fll=fill(NAVY), aln=align('right'))
    ws.row_dimensions[r].height = 15
    r += 1

    # ── SOLD TO / SHIP TO labels ──────────────────────────────────────────────
    ws.merge_cells(f'A{r}:B{r}')
    set_cell(ws, f'A{r}', 'SOLD TO',
             fnt=Font(bold=True, color=WHITE, size=8.5, name='Arial'),
             fll=fill('1A4A8A'), aln=align('center'))
    ws.merge_cells(f'C{r}:E{r}')
    set_cell(ws, f'C{r}', 'SHIP TO',
             fnt=Font(bold=True, color=WHITE, size=8.5, name='Arial'),
             fll=fill('1A4A8A'), aln=align('center'))
    ws.row_dimensions[r].height = 15
    r += 1

    # ── SOLD TO / SHIP TO content (3 rows merged) ─────────────────────────────
    import math as _math
    addr_parts = list(filter(None, [
        consignee.get('name', ''),
        consignee.get('address1', ''),
        consignee.get('address2', ''),
    ]))
    addr_lines = '\n'.join(addr_parts)
    # estimate lines needed in the narrower left column (A+B = ~27 char units)
    col_ab_chars = int(ws.column_dimensions['A'].width + ws.column_dimensions['B'].width)
    total_lines = sum(_math.ceil(len(p) / col_ab_chars) for p in addr_parts) if addr_parts else 1
    block_height = max(42, total_lines * 14)
    addr_end = r + 2
    ws.merge_cells(f'A{r}:B{addr_end}')
    set_cell(ws, f'A{r}', addr_lines,
             fnt=font(size=9), fll=fill(LT_GRAY),
             aln=align('left', 'top', wrap=True),
             brd=border(left=True, right=True, top=True, bottom=True))
    ws.merge_cells(f'C{r}:E{addr_end}')
    set_cell(ws, f'C{r}', addr_lines,
             fnt=font(size=9), fll=fill(LT_GRAY),
             aln=align('left', 'top', wrap=True),
             brd=border(left=True, right=True, top=True, bottom=True))
    per_row_h = block_height / 3
    for i in range(r, addr_end + 1):
        ws.row_dimensions[i].height = per_row_h
    r = addr_end + 1

    # ── Shipment info grid ────────────────────────────────────────────────────
    info_rows = [
        ('Ref Order',          reference,       'Port of Origin',    port_origin),
        ('Customer PO#',       po_number,       'Container #',       container_num),
        ('Seal #',             seal_num,        'Final Destination',  final_dest),
        ('Country of Origin',  'United States', 'Mode of Transport', mode_transport),
    ]
    for label1, val1, label2, val2 in info_rows:
        set_cell(ws, f'A{r}', label1,
                 fnt=font(bold=True, color=NAVY, size=8.5),
                 fll=fill(MED_GRAY), aln=align('left'),
                 brd=border(left=True, top=True, bottom=True))
        set_cell(ws, f'B{r}', val1,
                 fnt=font(size=9), fll=fill(WHITE), aln=align('left'),
                 brd=border(right=True, top=True, bottom=True))
        set_cell(ws, f'C{r}', label2,
                 fnt=font(bold=True, color=NAVY, size=8.5),
                 fll=fill(MED_GRAY), aln=align('left'),
                 brd=border(left=True, top=True, bottom=True))
        ws.merge_cells(f'D{r}:E{r}')
        set_cell(ws, f'D{r}', val2,
                 fnt=font(size=9), fll=fill(WHITE), aln=align('left'),
                 brd=border(left=True, right=True, top=True, bottom=True))
        ws.row_dimensions[r].height = 16
        r += 1

    # ── Table header ──────────────────────────────────────────────────────────
    headers = [
        ('Total QTY',        'center'),
        ('Package Type',     'center'),
        ('Total # Of Pieces','center'),
        ('Part Description', 'left'),
        ('Weight (LBS)',     'right'),
    ]
    for i, (h, ha) in enumerate(headers):
        col = get_column_letter(i + 1)
        set_cell(ws, f'{col}{r}', h,
                 fnt=Font(bold=True, color=WHITE, size=9, name='Arial'),
                 fll=fill(NAVY), aln=align(ha),
                 brd=border(left=(i == 0), right=(i == 4), top=True, bottom=True))
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Product rows ──────────────────────────────────────────────────────────
    for idx, it in enumerate(items):
        qty    = it.get('qty', 0)
        pres   = it.get('presentation', '')
        desc   = it.get('description', '')
        wt     = it.get('weight_lbs', 0)
        pieces = qty * PIECES_MULT.get(pres, 1)
        bg     = WHITE if idx % 2 == 0 else LT_GRAY

        row_vals = [
            (qty,         'center'),
            (pres,        'center'),
            (pieces,      'center'),
            (desc,        'left'),
            (round(wt),   'right'),
        ]
        for i, (v, ha) in enumerate(row_vals):
            col = get_column_letter(i + 1)
            set_cell(ws, f'{col}{r}', v,
                     fnt=font(size=9), fll=fill(bg), aln=align(ha),
                     brd=border(left=(i == 0), right=(i == 4), bottom=True))
        ws.row_dimensions[r].height = 15
        r += 1

    r += 1  # spacer

    # ── Summary rows ──────────────────────────────────────────────────────────
    net_wt   = round(sum(it.get('weight_lbs', 0) for it in items))
    gross_wt = round(net_wt + container_wt)

    for label, val, bg_color in [
        ('QUANTITY OF PALLETS',    pallets,             LT_GRAY),
        ('CONTAINER WEIGHT (LBS)', round(container_wt), SUM1_BG),
        ('TOTAL NET WEIGHT (LBS)', net_wt,              SUM2_BG),
        ('TOTAL GROSS WEIGHT (LBS)', gross_wt,          SUM3_BG),
    ]:
        ws.merge_cells(f'A{r}:D{r}')
        set_cell(ws, f'A{r}', label,
                 fnt=Font(bold=True, color=NAVY, size=9, name='Arial'),
                 fll=fill(bg_color), aln=align('right'),
                 brd=border(left=True, right=True, top=True, bottom=True))
        set_cell(ws, f'E{r}', val,
                 fnt=Font(bold=True, color=NAVY, size=10, name='Arial'),
                 fll=fill(bg_color), aln=align('right'),
                 brd=border(left=True, right=True, top=True, bottom=True))
        ws.row_dimensions[r].height = 17
        r += 1

    # ── Footer ────────────────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f'A{r}:C{r}')
    set_cell(ws, f'A{r}', f'Comments:  {comments}',
             fnt=font(italic=True, size=9, color='555555'), aln=align('left'))
    ws.row_dimensions[r].height = 16
    r += 1

    set_cell(ws, f'A{r}', f'Date: {date_display}', fnt=font(size=9))
    ws.merge_cells(f'C{r}:E{r}')
    set_cell(ws, f'C{r}', 'Authorized Signature:  _________________________________',
             fnt=font(size=9), aln=align('right'))
    ws.row_dimensions[r].height = 18

    if _logo_img:
        ws.add_image(_logo_img)

    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    out = io_mod.BytesIO()
    wb.save(out)
    out.seek(0)
    cn = consignee.get('name', 'EXPORT').replace('/', '-').replace('\\', '-')
    return send_file(out, as_attachment=True,
                     download_name=f"Packing List - {cn} - {reference}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── QuickBooks integration routes ─────────────────────────────────────────────

@app.route('/api/push-to-qb', methods=['POST'])
def push_to_qb():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'No JSON body received'}), 400
    job_type = data.get('job_type', 'sales_order')
    if job_type not in ('sales_order', 'estimate'):
        return jsonify({'error': 'job_type must be sales_order or estimate'}), 400
    customer = (data.get('customer_name') or '').strip()
    if not customer:
        return jsonify({'error': 'customer_name is required'}), 400
    line_items = data.get('line_items', [])
    if not line_items:
        return jsonify({'error': 'line_items cannot be empty'}), 400
    payload = {
        'customer_name': customer,
        'ref_number':    data.get('ref_number', ''),
        'memo':          data.get('memo', ''),
        'txn_date':      data.get('txn_date', ''),
        'line_items':    line_items,
    }
    job_id = enqueue(job_type, payload)
    return jsonify({'job_id': job_id, 'status': 'queued'})


@app.route('/api/qb-job-status/<job_id>', methods=['GET'])
def qb_job_status(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    resp = {
        'status':     job['status'],
        'created_at': job['created_at'],
        'updated_at': job['updated_at'],
    }
    if job['result']:
        resp['result'] = _json.loads(job['result'])
    if job['error']:
        resp['error'] = job['error']
    return jsonify(resp)


@app.route('/api/qb-inventory', methods=['POST'])
def qb_inventory():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'No JSON body received'}), 400
    item_codes = data.get('item_codes', [])
    if not item_codes:
        return jsonify({'error': 'item_codes cannot be empty'}), 400
    job_id = enqueue('inventory_query', {'item_codes': item_codes})
    deadline = time.time() + 30
    while time.time() < deadline:
        time.sleep(2)
        job = get_job(job_id)
        if not job:
            break
        if job['status'] == 'done':
            result = _json.loads(job['result'])
            return jsonify(result.get('data', {}))
        if job['status'] == 'error':
            return jsonify({'error': job['error']}), 502
    return jsonify({'error': 'QB is not responding. Is QBWC running?', 'job_id': job_id}), 504


@app.route('/api/qb-jobs', methods=['GET'])
def qb_jobs_list():
    jobs = get_recent_jobs(50)
    for j in jobs:
        if j.get('result'):
            j['result'] = _json.loads(j['result'])
    return jsonify(jobs)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if not (BASE / 'products.json').exists():
        print("WARNING: products.json not found. Run build_catalog.py first.")
    print("Starting U1P Order Conversion Tool -> http://localhost:5000")
    app.run(debug=False, port=5000)
